"""Redraw the sweep workbook's figures in matplotlib.

Reads an aggregate ``results.xlsx`` written by ``benchmarks/runner.py`` and
re-plots every chart it holds -- same data, same titles, same figure set --
in a hatched black-edged paper style instead of Excel's:

* one dual-axis latency (bars, seconds) + DRAM traffic (line, GB) figure per
  sweep group and mode;
* one stacked DRAM Weight / Activation / KV breakdown (GB) beside it;
* one stacked per-module compute / overlap / memory / stall breakdown
  (million cycles), prefill over decode, out of the two baseline sheets.

Nothing here imports the benchmarks package: the workbook carries the group
column and the chart titles, so the sweep definitions do not have to be
repeated.

    python plot_results_mpl.py
    python plot_results_mpl.py --workbook path/to/results.xlsx --out figures
    python plot_results_mpl.py --only context --only hardware --log
"""

import argparse
import difflib
import os
import re
from functools import partial

import matplotlib

matplotlib.use("Agg")

import matplotlib.patheffects as patheffects
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook

DEFAULT_WORKBOOK = "benchmarks/results/latest/results.xlsx"
DEFAULT_OUT = "benchmarks/figures"

# Both metrics divide by 1e9: latency cycles @ 1 GHz -> seconds, bytes -> GB.
# The per-module breakdown stays in cycles, shown in millions.
CYCLES_PER_SECOND = 1e9
BYTES_PER_GB = 1e9
CYCLES_PER_MCYCLE = 1e6

# -- the style ----------------------------------------------------------------
FIGSIZE = (10.0, 3.8)
BAR_WIDTH = 0.30
EDGE_WIDTH = 1.0
HEADER_PT = 20
TEXT_PT = 19
# A merged prefill-over-decode figure keeps a small gap between the stacked
# panels -- wider when they do not share an x-axis, so the upper one's labels
# clear the lower.
PAIR_HSPACE = 0.05
PAIR_HSPACE_UNSHARED = 0.1
# Headroom above the top mark, leaving the value labels somewhere to go.  The
# two axes of the latency+DRAM chart get *different* headroom on purpose: most
# sweeps here are DRAM-bound at a near-constant GB/s, so latency is DRAM
# traffic over a constant and the two series are proportional -- on two axes
# each normalized to its own max they would draw the same line.  Parking the
# bars in the lower band and the line up top keeps them apart, and keeps the
# blue labels away from the red ones.
HEADROOM = 1.35
LATENCY_HEADROOM = 1.45
DRAM_HEADROOM = 1.12
# A stack is labeled with its total and nothing above it, so it needs only
# enough room for that one label -- even tilted upright on a crowded axis.
STACK_HEADROOM = 1.18
LOG_HEADROOM = 2.2
LOG_FLOOR = 0.7

# A sweep's category labels are short and read upright.  The per-module
# breakdown's ~20 block names ("post attention layernorm") are not: they only
# fit tilted, and even then the figure has to widen and the names to shrink.
CROWDED_POINTS = 10
LONG_LABEL = 10
CROWDED_ROT = 40
CROWDED_PT = 14
WIDTH_PER_POINT = 0.8
MAX_WIDTH = 20.0

# Value labels sit below the axis fonts: at tick-label size they are wide
# enough to run into their neighbors.
VALUE_PT = 14
CROWDED_VALUE_PT = 12
# Gap under a value label, and the line-height multiple used to stack one
# label clear of another (both in points).
LABEL_PAD = 4
LINE_HEIGHT = 1.25
# A steep DRAM line runs straight through its own label; a white halo keeps
# the digits readable where it does.
HALO = [patheffects.withStroke(linewidth=2.5, foreground="white")]

# A legend beside the axes is charged for its width: a swatch wide enough to
# read the hatch, and no more, with the text close behind it (both in font
# widths, matplotlib's unit for these).
LEGEND_HANDLE = 1.2
LEGEND_PAD = 0.5

HATCHES = ("\\\\", "//", "xx", "..", "++", "oo")
# Distinct enough in grayscale print; the hatch carries the series anyway.
COLORS = ("#4F81BD", "#C0504D", "#9BBB59", "#8064A2", "#4BACC6", "#F79646")

# The stacked DRAM breakdown: (legend name, sheet column, stack order).
DRAM_SERIES = (
    ("Weight", "dram_weight"),
    ("Activation", "dram_activation"),
    ("KV Cache", "dram_kv"),
)
# The stacked per-module breakdown, in stack order under the compute series --
# which is named for the unit that runs it, and so differs by mode.
MODULE_COMPUTE = {
    "prefill": "Matrix Unit (64x64)",
    "decode": "Vector Unit (64 wide)",
}
MODULE_SERIES = (
    ("Overlap", "compute memory overlap"),
    ("Memory (64 GB/s)", "memory"),
    ("Stall", "stall"),
)

BASELINE_SHEETS = ("baseline_prefill", "baseline_decode")
MODES = ("prefill", "decode")


def _layout(labels):
    """``(figsize, label_pt, value_pt, crowded)`` for a category axis: the
    style as written, unless the categories are many or long."""
    crowded = (
        len(labels) > CROWDED_POINTS
        or max(len(str(l)) for l in labels) > LONG_LABEL
    )
    if not crowded:
        return FIGSIZE, TEXT_PT, VALUE_PT, False
    width = min(MAX_WIDTH, max(FIGSIZE[0], WIDTH_PER_POINT * len(labels) + 4))
    return (width, FIGSIZE[1] + 0.8), CROWDED_PT, CROWDED_VALUE_PT, True


def _bold_ticks(ax, labels, x, pt=TEXT_PT, crowded=False):
    """Upright bold category labels under ``x``, bold value ticks beside.  Only
    a crowded axis tilts them -- the per-module breakdown's block names are far
    too long to sit upright without running into each other."""
    rot = CROWDED_ROT if crowded else 0
    ax.set_xticks(x)
    ax.set_xticklabels(
        labels,
        rotation=rot,
        ha="right" if rot else "center",
        fontsize=pt,
        fontweight="bold",
    )
    ax.tick_params(axis="y", labelsize=HEADER_PT)
    for lbl in ax.get_yticklabels():
        lbl.set_fontweight("bold")


def _grid(ax):
    ax.set_axisbelow(True)
    ax.grid(axis="y", linestyle="--", linewidth=0.6, alpha=0.65, zorder=0)


def _fmt(v):
    """A value label: 3 significant digits, without a trailing ``.00``."""
    if v == 0:
        return "0"
    if abs(v) >= 100:
        return f"{v:,.0f}"
    if abs(v) >= 10:
        return f"{v:.1f}"
    if abs(v) >= 1:
        return f"{v:.2f}"
    return f"{v:.3g}"


def _axes_height_points(fig, ax):
    """The axes' drawn height in points -- the unit an offset label uses.  The
    layout has to run first, or the axes still hold their pre-layout size."""
    fig.canvas.draw()
    return ax.get_window_extent().height * 72.0 / fig.dpi


def _frac(ax, value):
    """Where ``value`` lands up ``ax``, as a fraction of the axes height."""
    lo, hi = ax.get_ylim()
    if ax.get_yscale() == "log":
        return (np.log10(value / lo)) / np.log10(hi / lo)
    return (value - lo) / (hi - lo)


def _label_bars(ax, bars, values, pt, color, rot=0):
    """Value labels above ``bars``, in the bars' own color.  A bar of no height
    is left unlabeled: there is no bar there to name."""
    ax.bar_label(
        bars,
        labels=["" if v == 0 else _fmt(v) for v in values],
        padding=4,
        fontsize=pt,
        fontweight="bold",
        color=color,
        rotation=rot,
        path_effects=HALO,
    )


def _legend(fig, handles, labels, ncol):
    """A frameless legend in its own band under the figure -- with the values
    labeled there is no room left inside the axes."""
    legend = fig.legend(
        handles,
        labels,
        loc="outside lower center",
        frameon=False,
        fontsize=HEADER_PT,
        ncol=ncol,
    )
    for text in legend.get_texts():
        text.set_fontweight("bold")
    return legend


def _limits(ax, values, log, headroom=HEADROOM):
    """0-based (or decade-based, under ``--log``) axis with legend headroom."""
    top = max(values) or 1.0
    if log:
        ax.set_yscale("log")
        positive = [v for v in values if v > 0]
        ax.set_ylim(min(positive or [top]) * LOG_FLOOR, top * LOG_HEADROOM)
    else:
        ax.set_ylim(0, top * headroom)


def _save(fig, out_dir, stem):
    """Write ``stem`` as a high-DPI PNG."""
    path = os.path.join(out_dir, f"{stem}.png")
    fig.savefig(path, dpi=600, bbox_inches="tight")
    plt.close(fig)
    return [path]


def _axis_latency_dram(fig, ax, labels, latency, dram, title, log, xlabel=None):
    """Draw latency (bars, left/seconds) + total DRAM traffic (line, right/GB)
    onto ``ax`` and its twinx.  Returns ``(handles, names)`` for a shared
    legend; the caller owns the figure, so this neither legends nor saves."""
    x = np.arange(len(labels))
    seconds = np.array(latency, dtype=float) / CYCLES_PER_SECOND
    gigabytes = np.array(dram, dtype=float) / BYTES_PER_GB
    _, pt, vpt, crowded = _layout(labels)

    _grid(ax)
    bars = ax.bar(
        x,
        seconds,
        BAR_WIDTH * 2,
        label="Latency",
        color=COLORS[0],
        hatch=HATCHES[0],
        edgecolor="black",
        linewidth=EDGE_WIDTH,
        zorder=3,
    )
    ax.set_ylabel("Latency (s)", fontsize=HEADER_PT, fontweight="bold")
    _limits(ax, list(seconds), log, headroom=LATENCY_HEADROOM)
    _bold_ticks(ax, labels, x, pt, crowded)
    if xlabel:
        ax.set_xlabel(xlabel, fontsize=HEADER_PT, fontweight="bold")
    _label_bars(ax, bars, seconds, vpt, COLORS[0])

    right = ax.twinx()
    line = right.plot(
        x,
        gigabytes,
        label="DRAM Traffic",
        color=COLORS[1],
        linewidth=2.5,
        marker="o",
        markersize=7,
        markeredgecolor="black",
        markeredgewidth=0.8,
        zorder=4,
    )[0]
    right.set_ylabel("DRAM (GB)", fontsize=HEADER_PT, fontweight="bold")
    _limits(right, list(gigabytes), log, headroom=DRAM_HEADROOM)
    right.tick_params(axis="y", labelsize=HEADER_PT)
    for lbl in right.get_yticklabels():
        lbl.set_fontweight("bold")
    # The DRAM label goes above its marker -- but where the two series are
    # proportional the marker sits right on the bar top, so lift it clear of
    # the latency label instead of drawing one on top of the other.
    axes_pt = _axes_height_points(fig, ax)
    for xi, gb, sec in zip(x, gigabytes, seconds):
        bar_pt = axes_pt * _frac(ax, sec)
        blue_top = bar_pt + LABEL_PAD + LINE_HEIGHT * vpt
        marker_pt = axes_pt * _frac(right, gb)
        rise = max(LABEL_PAD + 2, blue_top + LABEL_PAD - marker_pt)
        right.annotate(
            _fmt(gb),
            (xi, gb),
            textcoords="offset points",
            xytext=(0, rise),
            ha="center",
            fontsize=vpt,
            fontweight="bold",
            color=COLORS[1],
            path_effects=HALO,
        )
    ax.set_title(title, fontsize=HEADER_PT, fontweight="bold")
    return [bars, line], ["Latency", "DRAM Traffic"]


def _axis_stacked(fig, ax, labels, series, ylabel, title, log, xlabel=None):
    """Draw a stacked bar per category onto ``ax``; ``series`` is ordered
    ``(name, values)``.  Returns ``(handles, names)`` for a shared legend."""
    x = np.arange(len(labels))
    _, pt, vpt, crowded = _layout(labels)
    _grid(ax)

    bottom = np.zeros(len(labels))
    handles = []
    for i, (name, values) in enumerate(series):
        values = np.array(values, dtype=float)
        bars = ax.bar(
            x,
            values,
            BAR_WIDTH * 2,
            bottom=bottom,
            label=name,
            color=COLORS[i % len(COLORS)],
            hatch=HATCHES[i % len(HATCHES)],
            edgecolor="black",
            linewidth=EDGE_WIDTH,
            zorder=3,
        )
        bottom = bottom + values
        handles.append(bars)

    ax.set_ylabel(ylabel, fontsize=HEADER_PT, fontweight="bold")
    # Only the stack total is labeled: a segment's own value would land inside
    # the stack, on top of the hatch, and the thin ones have nowhere to go.
    # ``bars`` is the topmost segment, so a label above it is a label above the
    # stack; on a crowded axis the totals turn upright to stay apart.
    _limits(ax, list(bottom), log, headroom=STACK_HEADROOM)
    _bold_ticks(ax, labels, x, pt, crowded)
    if xlabel:
        ax.set_xlabel(xlabel, fontsize=HEADER_PT, fontweight="bold")
    _label_bars(ax, bars, bottom, vpt, "black", rot=90 if crowded else 0)
    ax.set_title(title, fontsize=HEADER_PT, fontweight="bold")
    return handles, [n for n, _ in series]


def draw_pair(
    calls,
    suptitle,
    base_figsize,
    out_dir,
    stem,
    sharex=True,
    legend_right=False,
):
    """One figure with a subplot per entry of ``calls`` stacked top to bottom --
    prefill above decode -- under one shared ``suptitle``.  Each ``call(fig,
    ax)`` draws one mode, titles its panel with the mode, and returns
    ``(handles, names)`` for the shared legend.  With ``sharex`` the panels
    plot the same points and share one x-axis, labeled once at the bottom;
    without it each panel carries its own tick + axis labels (the pe sweep,
    whose modes read different axes).  ``legend_right`` gives each panel its
    own legend, beside it, instead of one shared band under the figure -- for
    a pair whose series are named differently per mode."""
    w, h = base_figsize
    fig, axes = plt.subplots(
        len(calls),
        1,
        figsize=(w, h * len(calls)),
        layout="constrained",
        sharex=sharex,
    )
    fig.get_layout_engine().set(
        hspace=PAIR_HSPACE if sharex else PAIR_HSPACE_UNSHARED
    )
    axes = np.atleast_1d(axes)
    handles = names = None
    for ax, call in zip(axes, calls):
        handles, names = call(fig, ax)
        if legend_right:
            legend = ax.legend(
                handles,
                names,
                loc="center left",
                bbox_to_anchor=(1.01, 0.5),
                frameon=False,
                fontsize=HEADER_PT,
                handlelength=LEGEND_HANDLE,
                handletextpad=LEGEND_PAD,
                borderaxespad=0,
            )
            for text in legend.get_texts():
                text.set_fontweight("bold")
        # A shared x-axis is labeled once, at the bottom.
        if sharex:
            ax.label_outer()
    fig.suptitle(suptitle, fontsize=HEADER_PT, fontweight="bold")
    if not legend_right:
        _legend(fig, handles, names, ncol=len(names))
    return _save(fig, out_dir, stem)


def _latency_axis(fig, ax, entry, log):
    """Adapt a metric-sheet ``(block, mode, lat, brk, labels, xlabel)`` entry to
    :func:`_axis_latency_dram` -- one ``draw_pair`` panel, titled by its mode.
    """
    block, mode, _lat, _brk, labels, xlabel = entry
    return _axis_latency_dram(
        fig,
        ax,
        labels,
        [r["total_latency"] for r in block],
        [r["dram_total"] for r in block],
        mode.capitalize(),
        log,
        xlabel,
    )


def _breakdown_axis(fig, ax, entry, log):
    """Adapt a metric-sheet entry to the DRAM Weight/Activation/KV stack."""
    block, mode, _lat, _brk, labels, xlabel = entry
    series = [
        (name, [(r[field] or 0) / BYTES_PER_GB for r in block])
        for name, field in DRAM_SERIES
    ]
    return _axis_stacked(
        fig, ax, labels, series, "DRAM (GB)", mode.capitalize(), log, xlabel
    )


def _rows(ws):
    """A sheet's populated rows as dicts keyed by its header, blanks dropped."""
    header, *body = ws.values
    keys = [h for h in header]
    out = []
    for row in body:
        if row[0] is None and row[1] is None:
            continue  # the blank separator row between groups
        out.append(dict(zip(keys, row)))
    return out


def _chart_titles(ws):
    """The sheet's chart titles, in the order the charts were added."""
    titles = []
    for chart in ws._charts:
        rich = chart.title.tx.rich
        titles.append("".join(r.t or "" for p in rich.p for r in (p.r or [])))
    return titles


def _stem(*parts):
    """A filesystem-safe figure name out of the sheet / group / mode parts."""
    text = "_".join(str(p) for p in parts if p)
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", text).strip("_").lower()


def _shared_title(title):
    """The two panels of a pair are titled ``Prefill``/``Decode``; their common
    text -- the workbook title with the mode word dropped -- is the figure's
    single suptitle (e.g. ``Llama 3.1 8B Prefill Latency ...`` ->
    ``Llama 3.1 8B Latency ...``)."""
    for mode in MODES:
        title = title.replace(f"{mode.capitalize()} ", "")
    return title


def _axis_spec(group, mode, points):
    """A ``(panel, mode)``'s ``(tick labels, x-axis label)``.  Only the hardware
    ``pe`` sweep relabels: prefill fills the whole NxN array (``16x16``, "PE
    Array Size"), decode is vector-bound and reads one dimension as lanes (the
    trailing number, "Vector Lanes").  Every other sweep keeps its points,
    unlabeled and shared across the two panels."""
    if group == "pe":
        if mode == "decode":
            return [p.split("x")[-1] for p in points], "Vector Lanes"
        return list(points), "PE Array Size"
    return list(points), None


def plot_metric_sheet(ws, log, out_dir):
    """One sweep sheet: per group, a latency+DRAM figure and a DRAM-breakdown
    figure, each with prefill above decode under a single shared title (the
    workbook's, minus the per-mode word)."""
    rows = _rows(ws)
    titles = _chart_titles(ws)
    written = []

    # Charts were added group-major, prefill before decode, latency then
    # breakdown.  Walk that order to pair each (group, mode) with its two
    # titles, collecting the modes of a group side by side.
    grouped = (
        {}
    )  # group -> [entry]; entry = (block, mode, lat, brk, labels, xl)
    idx = 0
    for group in dict.fromkeys(r["group"] for r in rows):
        for mode in MODES:
            block = [
                r for r in rows if r["group"] == group and r["mode"] == mode
            ]
            if not block:
                continue
            lat = titles[idx] if idx < len(titles) else ws.title
            brk = titles[idx + 1] if idx + 1 < len(titles) else ws.title
            idx += 2
            labels, xlabel = _axis_spec(
                group, mode, [r["point"] for r in block]
            )
            grouped.setdefault(group, []).append(
                (block, mode, lat, brk, labels, xlabel)
            )

    for group, entries in grouped.items():
        base = _layout(entries[0][4])[0]
        # Share one x-axis only when both panels label it the same way.
        sharex = len({(tuple(e[4]), e[5]) for e in entries}) == 1
        written += draw_pair(
            [partial(_latency_axis, entry=e, log=log) for e in entries],
            _shared_title(entries[0][2]),
            base,
            out_dir,
            _stem(ws.title, group, "latency_dram"),
            sharex,
        )
        written += draw_pair(
            [partial(_breakdown_axis, entry=e, log=log) for e in entries],
            _shared_title(entries[0][3]),
            base,
            out_dir,
            _stem(ws.title, group, "dram_breakdown"),
            sharex,
        )
    return written


def _module_key(block):
    """A block name as it compares across modes: the decode graph nests the
    model one module deeper, and a fused block carries the suffix."""
    stem = block.removeprefix("model_").removeprefix("model_")
    return stem.removesuffix("_fused")


def _align_modules(rows_per_mode):
    """One shared module axis for the modes' block lists.  The two graphs are
    near-identical but not equal -- decode runs the cache's quantize ops and an
    extra attention matmul -- so they are merged the way a diff would: blocks
    that match share a slot, and a block only one mode runs takes a slot of its
    own, left empty in the other panel.  Returns the slot labels and, per mode,
    the row filling each slot (``None`` where it has none)."""
    if len(rows_per_mode) == 1:
        return [r["display"] for r in rows_per_mode[0]], rows_per_mode
    left, right = rows_per_mode
    diff = difflib.SequenceMatcher(
        None,
        [_module_key(r["block"]) for r in left],
        [_module_key(r["block"]) for r in right],
        autojunk=False,
    )
    labels, slots = [], ([], [])
    for tag, i1, i2, j1, j2 in diff.get_opcodes():
        if tag == "equal":
            for i, j in zip(range(i1, i2), range(j1, j2)):
                labels.append(left[i]["display"])
                slots[0].append(left[i])
                slots[1].append(right[j])
            continue
        # A block one mode does not run: its own slot, empty in the other.
        for side, rows, lo, hi in ((0, left, i1, i2), (1, right, j1, j2)):
            for k in range(lo, hi):
                labels.append(rows[k]["display"])
                slots[side].append(rows[k])
                slots[1 - side].append(None)
    return labels, list(slots)


def _module_axis(fig, ax, labels, rows, mode, log):
    """Adapt a baseline sheet's rows to the per-module compute / overlap /
    memory / stall stack -- one ``draw_pair`` panel, titled by its mode.  A
    ``None`` row is a module the other mode runs and this one does not, and
    stacks to nothing."""
    series = [
        (
            name,
            [
                0.0 if r is None else (r[field] or 0) / CYCLES_PER_MCYCLE
                for r in rows
            ],
        )
        for name, field in ((MODULE_COMPUTE[mode], "compute"),) + MODULE_SERIES
    ]
    return _axis_stacked(
        fig,
        ax,
        labels,
        series,
        "Million cycles",
        mode.capitalize(),
        log,
    )


def plot_baseline_pair(sheets, log, out_dir):
    """The baseline sheets' per-module runtime breakdowns as one figure,
    prefill above decode under a single shared title, over one shared module
    axis labeled once at the bottom.  Each panel legends itself, since the
    compute series is named for the unit that ran it."""
    modes = [ws.title.split("_")[-1] for ws in sheets]
    labels, aligned = _align_modules([_rows(ws) for ws in sheets])
    titles = _chart_titles(sheets[0])
    return draw_pair(
        [
            partial(_module_axis, labels=labels, rows=rows, mode=mode, log=log)
            for mode, rows in zip(modes, aligned)
        ],
        _shared_title(titles[0] if titles else sheets[0].title),
        _layout(labels)[0],
        out_dir,
        _stem("baseline", "breakdown"),
        legend_right=True,
    )


def parse_args():
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument(
        "--workbook",
        default=DEFAULT_WORKBOOK,
        help="Aggregate results.xlsx to re-plot.",
    )
    p.add_argument("--out", default=DEFAULT_OUT, help="Figure output dir.")
    p.add_argument(
        "--only",
        action="append",
        help="Restrict to these sheet names (repeatable); default: all.",
    )
    p.add_argument(
        "--log",
        action="store_true",
        help="Log-scale the value axes (sweeps spanning decades).",
    )
    return p.parse_args()


def main():
    args = parse_args()
    wb = load_workbook(args.workbook)
    os.makedirs(args.out, exist_ok=True)

    names = args.only or wb.sheetnames
    written = []
    for name in names:
        if name not in BASELINE_SHEETS:
            written += plot_metric_sheet(wb[name], args.log, args.out)
    baselines = [wb[n] for n in BASELINE_SHEETS if n in names]
    if baselines:
        written += plot_baseline_pair(baselines, args.log, args.out)

    for path in written:
        print(path)
    print(f"\nwrote {len(written)} files to {os.path.abspath(args.out)}")


if __name__ == "__main__":
    main()
