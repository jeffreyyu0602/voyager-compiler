"""Styled per-module runtime-breakdown chart for the baseline sweep.

The baseline figure: one stacked bar per compute block, its runtime split into
compute / overlap / memory / stall, each scaled by the block's whole-model
repeat count.  Styled like the hand-built reference
(``baseline_prefill_chart_updated.xlsx``): legacy Office-theme colors, Calibri,
a bottom legend, light gridlines, and a value axis scaled to millions.

``populate_sheet`` writes the data + chart into an existing worksheet from the
sweep's ``Metrics`` records -- how ``runner.py`` drops the baseline breakdown
into the aggregate (via ``latency_dram_chart``).  ``write_per_module_chart``
wraps it into a standalone workbook, and ``restyle`` (+ the CLI) re-styles an
already-written plain breakdown workbook.
"""

import sys

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.drawing.text import (
    CharacterProperties,
    Font,
    Paragraph,
    ParagraphProperties,
    RichTextProperties,
)

# The hand-built workbook's palette, read straight out of its chart XML: the
# series fills resolve through the legacy Office theme
# (accent1 / tx2 / accent4), so they are spelled here as literal RGB.
COMPUTE = "4F81BD"  # accent1  -- steel blue
OVERLAP = "1F497D"  # tx2      -- dark navy
MEMORY = "C00000"  # explicit  -- dark red
STALL = "8064A2"  # accent4  -- purple
GRID = "D9D9D9"  # tx1 @ 15% luminance -- the gridline / axis gray

# Column order of the plotted series (fill + header label per column).
SERIES_FILL = {
    "compute": COMPUTE,
    "overlap": OVERLAP,
    "memory": MEMORY,
    "stall": STALL,
}

# Series names come from the header row, so a legend entry is renamed by
# renaming the column it reads.
HEADER_LABEL = {"overlap": "compute memory overlap"}

# The four plotted series occupy columns C..F (so the chart's column
# references stay put); the rest of the Metrics fields are written after them,
# as ``(header, attribute)`` pairs, to keep the whole per-block record.
EXTRA_COLS = (
    ("cycles", "total_latency"),
    ("dram_read", "dram_read_bytes"),
    ("dram_write", "dram_write_bytes"),
    ("dram_total", "dram_total_bytes"),
    ("dram_weight", "dram_weight_bytes"),
    ("dram_activation", "dram_activation_bytes"),
    ("dram_kv", "dram_kv_bytes"),
    ("scratchpad", "scratchpad_bytes"),
    ("count", "count"),
)

FONT = "Calibri"
TITLE_PT = 14
CAT_PT = 10
VAL_PT = 9
VAL_UNIT_PT = 10
LEGEND_PT = 10

# Excel's "auto rotation" sentinel; copied verbatim from the reference so the
# tick labels lay out identically.
AUTO_ROT = -60000000


def _char(size_pt):
    """Calibri at ``size_pt``, black, plain."""
    return CharacterProperties(
        latin=Font(typeface=FONT),
        sz=size_pt * 100,
        b=False,
        i=False,
        u="none",
        strike="noStrike",
        kern=1200,
        baseline=0,
    )


def _text(size_pt, rot=AUTO_ROT):
    """A RichText run-property block: Calibri at ``size_pt``, black, upright."""
    props = _char(size_pt)
    body = RichTextProperties(
        rot=rot,
        spcFirstLastPara=1,
        vertOverflow="ellipsis",
        vert="horz",
        wrap="square",
        anchor="ctr",
        anchorCtr=1,
    )
    para = Paragraph(pPr=ParagraphProperties(defRPr=props), endParaRPr=props)
    return RichText(bodyPr=body, p=[para])


def build_chart(ws, title, n_rows, first_row=2):
    """A stacked column chart over ``ws`` rows ``first_row..first_row+n_rows``,
    styled like the reference workbook."""
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.gapWidth = 150
    chart.roundedCorners = False
    chart.title = title
    chart.dLbls = DataLabelList(
        showLegendKey=False,
        showVal=False,
        showCatName=False,
        showSerName=False,
        showPercent=False,
        showBubbleSize=False,
    )

    last = first_row + n_rows - 1
    # Categories are text, so they must be a strRef -- openpyxl's
    # set_categories would emit a numRef and Excel would show 1..N instead.
    cats = AxDataSource(strRef=StrRef(f=f"{ws.title}!$B${first_row}:$B${last}"))
    for i, name in enumerate(SERIES_FILL):
        col = 3 + i
        ref = Reference(ws, min_col=col, min_row=1, max_row=last)
        s = Series(ref, title_from_data=True)
        s.graphicalProperties = GraphicalProperties(
            solidFill=SERIES_FILL[name], ln=LineProperties(noFill=True)
        )
        s.cat = cats
        chart.series.append(s)

    # Title: 14pt Calibri, upright (no auto-rotation sentinel).  ``overlay``
    # must be off or Excel floats the title over the plot instead of
    # reserving a band for it.
    chart.title.overlay = False
    chart.title.tx.rich.p[0].pPr = ParagraphProperties(
        defRPr=_text(TITLE_PT, rot=0).p[0].pPr.defRPr
    )
    chart.title.txPr = _text(TITLE_PT, rot=0)
    chart.title.graphicalProperties = GraphicalProperties(
        noFill=True, ln=LineProperties(noFill=True)
    )

    # Category axis: hairline gray baseline, no ticks.
    x = chart.x_axis
    x.delete = False
    x.majorTickMark = "none"
    x.minorTickMark = "none"
    x.tickLblPos = "nextTo"
    x.spPr = GraphicalProperties(
        noFill=True,
        ln=LineProperties(
            solidFill=GRID, w=9525, cap="flat", cmpd="sng", algn="ctr"
        ),
    )
    x.txPr = _text(CAT_PT)

    # Value axis: no spine, light gridlines, scaled to millions.
    y = chart.y_axis
    y.delete = False
    y.majorTickMark = "none"
    y.minorTickMark = "none"
    y.tickLblPos = "nextTo"
    y.majorGridlines = ChartLines(
        spPr=GraphicalProperties(
            ln=LineProperties(
                solidFill=GRID, w=9525, cap="flat", cmpd="sng", algn="ctr"
            )
        )
    )
    # Scale to millions in the tick format itself (the two trailing commas),
    # not via dispUnits: Excel applies dispUnits to a series' data labels too,
    # which diverges from LibreOffice; a plain format renders identically in
    # both, and the unit lives in an ordinary axis title.
    y.numFmt = "#,##0,,"
    y.spPr = GraphicalProperties(noFill=True, ln=LineProperties(noFill=True))
    y.txPr = _text(VAL_PT)
    y.title = "Million cycles"
    y.title.tx.rich.p[0].pPr = ParagraphProperties(defRPr=_char(VAL_UNIT_PT))
    y.title.tx.rich.p[0].r[0].rPr = _char(VAL_UNIT_PT)
    y.title.overlay = False

    chart.legend.position = "b"
    chart.legend.overlay = False
    chart.legend.txPr = _text(LEGEND_PT, rot=0)
    chart.legend.spPr = GraphicalProperties(
        noFill=True, ln=LineProperties(noFill=True)
    )

    # White chart area, no border.
    chart.graphical_properties = GraphicalProperties(
        solidFill="FFFFFF", ln=LineProperties(noFill=True)
    )
    return chart


def _add_chart(ws, title, n_rows):
    """Widen the label columns and drop the styled chart onto ``ws`` to the
    right of the data (does not save)."""
    for col, width in (("A", 33), ("B", 16.83), ("C", 10.16), ("D", 10.16)):
        ws.column_dimensions[col].width = width
    chart = build_chart(ws, title, n_rows)
    # Data now runs A..O (15 cols), so the chart sits to its right: Q9 .. Z23,
    # spanning 9 cols x 14 rows (the reference's proportions).
    chart.anchor = TwoCellAnchor(
        _from=AnchorMarker(col=16, colOff=99768, row=8, rowOff=7103),
        to=AnchorMarker(col=25, colOff=462078, row=22, rowOff=66530),
    )
    ws.add_chart(chart)


def populate_sheet(ws, rows, title, display_name):
    """Write the breakdown data + chart into an existing worksheet ``ws`` from
    the sweep's ``Metrics`` records (no save).  The records already carry their
    whole-model contribution (``report_per_module`` scales each block by its
    repeat ``count``), so the values are written as-is; the four series land in
    columns C..F for the chart, the remaining Metrics fields follow.
    ``display_name`` shortens the raw block name for the category label."""
    series = tuple(SERIES_FILL)
    ws.append(
        [
            "block",
            "display",
            *(HEADER_LABEL.get(s, s) for s in series),
            *(h for h, _ in EXTRA_COLS),
        ]
    )
    for r in rows:
        ws.append(
            [
                r.name,
                display_name(r.name),
                *(getattr(r, s) for s in series),
                *(getattr(r, a) for _, a in EXTRA_COLS),
            ]
        )
    _add_chart(ws, title, len(rows))


def write_per_module_chart(rows, path, title, display_name):
    """Write the styled breakdown workbook to ``path`` straight from the sweep's
    ``Metrics`` records."""
    wb = Workbook()
    ws = wb.active
    ws.title = "breakdown"
    populate_sheet(ws, rows, title, display_name)
    wb.save(path)
    return path


def restyle(src, dst, title):
    """Re-style an already-written plain breakdown workbook (``block / display /
    compute / overlap / memory / stall``) into ``dst``."""
    # Re-emit the data into a *fresh* workbook: loading the source in place
    # would carry its original (unstyled) chart along into the output.
    header, *body = load_workbook(src)["breakdown"].values
    wb = Workbook()
    ws = wb.active
    ws.title = "breakdown"
    ws.append([HEADER_LABEL.get(h, h) for h in header])
    for row in body:
        ws.append(list(row))
    _add_chart(ws, title, len(body))
    wb.save(dst)
    return dst


if __name__ == "__main__":
    src, dst, title = sys.argv[1], sys.argv[2], sys.argv[3]
    print(restyle(src, dst, title))
