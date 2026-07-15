"""Aggregate design-space workbook: one styled sheet per sweep.

Every whole-graph sweep produces the same two metrics per design point -- total
latency and total DRAM traffic -- so they all get the same figure: a dual-axis
chart with latency as blue columns on the left (scaled to seconds) and DRAM as a
red line on the right (scaled to GB), drawn once per prefill and once per decode
over the sweep's config points.  A sweep with sub-groups (hardware: pe / sram /
bw) gets one such pair per group, stacked down the sheet.

``write_aggregate`` collects these sheets -- plus the per-module ``baseline``
breakdown sheets (delegated to :mod:`per_module_latency_chart`) -- into a single
workbook, one named sheet each.

The seconds/GB scaling assumes a 1 GHz clock (cycles / 1e9 = seconds); no sweep
here varies the frequency.
"""

import math
import re
import zipfile
from collections import namedtuple

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.chart.axis import (
    ChartLines,
    DisplayUnitsLabel,
    DisplayUnitsLabelList,
)
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import Marker
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText, Text
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.drawing.text import (
    CharacterProperties,
    Font,
    Paragraph,
    ParagraphProperties,
    RegularTextRun,
    RichTextProperties,
)

import per_module_latency_chart

# ---- what a caller hands us -------------------------------------------------
# A metric sheet is one sweep; each of its groups becomes a prefill+decode chart
# pair.  ``rows`` are dicts keyed by the METRIC_FIELDS below plus point / mode.
MetricSheet = namedtuple("MetricSheet", "name prefix groups")
MetricGroup = namedtuple("MetricGroup", "axis_title rows")
# A breakdown sheet is the per-module baseline (its own stacked-bar chart).
BreakdownSheet = namedtuple("BreakdownSheet", "name title rows display_name")

# Uniform column layout of a metric sheet.  total_latency / dram_total sit at
# fixed columns so the chart can address them regardless of the sweep.
METRIC_FIELDS = (
    "total_latency",
    "dram_total",
    "dram_read",
    "dram_write",
    "dram_weight",
    "dram_activation",
    "dram_kv",
    "scratchpad",
    "num_layers",
    "num_params",
)
HEADER = ("group", "point", "mode", *METRIC_FIELDS)
POINT_COL = 2  # B: the config point (category axis)
LATENCY_COL = 4  # D: total_latency (bars)
DRAM_COL = 5  # E: dram_total (line)

# Legacy Office theme, matching the hand-built reference: latency blue,
# DRAM red.
LATENCY = "4F81BD"  # accent1
DRAM = "C0504D"  # accent2
GRID = "D9D9D9"  # tx1 @ 15% luminance -- gridline / axis gray

# DRAM-traffic breakdown stacked bar: (legend name, metric field, fill), in
# stack order, matching the reference's chart5/6 (weight / activation / kv).
BREAKDOWN_SERIES = (
    ("Weight", "dram_weight", "C0504D"),  # accent2 red
    ("Activation", "dram_activation", "9BBB59"),  # accent3 green
    ("KV Cache", "dram_kv", "4F81BD"),  # accent1 blue
)
BREAKDOWN_GAP = 219  # matches the reference

FONT = "Calibri"
TITLE_PT = 14
# Everything but the title reads at 12pt.
CAT_PT = 12
VAL_PT = 12
UNIT_PT = 12
LEGEND_PT = 12
LABEL_PT = 12

AUTO_ROT = -60000000  # Excel "auto rotation" sentinel for tick labels
VERT_ROT = -5400000  # -90 deg: the axis unit label reads up the axis

# Both metrics divide by 1e9: latency cycles @ 1 GHz -> seconds, bytes -> GB.
DISP_UNIT = "billions"
DISP_SCALE = 1e9
LABEL_FMT = "#,##0.00"


def _char(size_pt, color=None):
    """Calibri at ``size_pt``; black unless ``color`` (an RGB hex) is given."""
    return CharacterProperties(
        latin=Font(typeface=FONT),
        sz=size_pt * 100,
        b=False,
        i=False,
        u="none",
        strike="noStrike",
        kern=1200,
        baseline=0,
        solidFill=color,
    )


def _text(size_pt, rot=AUTO_ROT, color=None):
    """A RichText run-property block: Calibri at ``size_pt``."""
    props = _char(size_pt, color)
    body = RichTextProperties(
        rot=rot,
        spcFirstLastPara=1,
        vertOverflow="ellipsis",
        vert="horz",
        wrap="square",
        anchor="ctr",
        anchorCtr=1,
    )
    para = Paragraph(pPr=ParagraphProperties(defRPr=props), endParaRPr=props)
    return RichText(bodyPr=body, p=[para])


def _unit_axis(title, disp_pt=UNIT_PT):
    """A value axis scaled to billions, its unit named in a rotated dispUnits
    label (``Seconds`` / ``GB``)."""
    lbl = DisplayUnitsLabel(
        tx=Text(
            rich=RichText(
                bodyPr=RichTextProperties(
                    rot=VERT_ROT,
                    spcFirstLastPara=1,
                    vertOverflow="ellipsis",
                    vert="horz",
                    wrap="square",
                    anchor="ctr",
                    anchorCtr=1,
                ),
                p=[
                    Paragraph(
                        pPr=ParagraphProperties(defRPr=_char(disp_pt)),
                        r=[RegularTextRun(rPr=_char(disp_pt), t=title)],
                    )
                ],
            )
        ),
        spPr=GraphicalProperties(noFill=True, ln=LineProperties(noFill=True)),
    )
    return DisplayUnitsLabelList(builtInUnit=DISP_UNIT, dispUnitsLbl=lbl)


def _labels(color, pos=None):
    """Value labels at ``#,##0.00`` in ``color``.  The format is stamped
    ``sourceLinked=0`` after save so a reader honors it."""
    return DataLabelList(
        dLblPos=pos,
        numFmt=LABEL_FMT,
        showVal=True,
        showLegendKey=False,
        showCatName=False,
        showSerName=False,
        showPercent=False,
        showBubbleSize=False,
        txPr=_text(LABEL_PT, rot=0, color=color),
        spPr=GraphicalProperties(noFill=True, ln=LineProperties(noFill=True)),
    )


def _col(i):
    return chr(ord("A") + i - 1)


def _field_col(field):
    """1-based sheet column of a METRIC_FIELDS value (past group/point/mode)."""
    return 4 + METRIC_FIELDS.index(field)


def _bind(s, sheet, value_col, start, end):
    """Give series ``s`` its name (header cell) and category (col B), over
    sheet rows ``start..end``."""
    s.tx = SeriesLabel(strRef=StrRef(f=f"{sheet}!${_col(value_col)}$1"))
    s.cat = AxDataSource(
        strRef=StrRef(
            f=f"{sheet}!${_col(POINT_COL)}${start}:" f"${_col(POINT_COL)}${end}"
        )
    )
    return s


# The dispUnits scale a tick's displayed value; the axis min/max/majorUnit stay
# in native units, so pick the "nice" step on the scaled value and multiply
# back.
NICE_STEPS = (1, 2, 2.5, 5, 10)
TARGET_TICKS = 5
# Fraction of each axis kept free above the top data point.  Both series label
# above their mark, but the latency (bar) axis gets far more headroom, parking
# the bars and their blue labels in the lower band while the DRAM line and its
# red labels stay up top -- so the two label sets do not collide.
DRAM_HEADROOM = 0.15
LATENCY_HEADROOM = 0.4


def nice_axis(data_max, headroom=DRAM_HEADROOM):
    """For a value axis 0..``data_max`` (native units) return ``(axis_max,
    major_unit, num_fmt)``: ~5 evenly spaced ticks at a fixed number of decimals
    with ``headroom`` above the top point.  Computed on the dispUnits-scaled
    value so the displayed ticks (seconds / GB) are round."""
    scaled = (data_max or 0) / DISP_SCALE
    if scaled <= 0:
        return DISP_SCALE, DISP_SCALE / TARGET_TICKS, "0.0"
    raw = scaled / TARGET_TICKS
    mag = 10 ** math.floor(math.log10(raw))
    step = next((m * mag for m in NICE_STEPS if m * mag >= raw), 10 * mag)
    axis_max = math.ceil(scaled / step) * step
    while axis_max - scaled < headroom * axis_max:
        axis_max += step
    d = max(0, -math.floor(math.log10(step)))
    while abs(step * 10**d - round(step * 10**d)) > 1e-9:
        d += 1
    fmt = "0." + "0" * d if d else "0"
    return axis_max * DISP_SCALE, step * DISP_SCALE, fmt


def _col_max(ws, col, start, end):
    return max(
        ws.cell(row=r, column=col).value or 0 for r in range(start, end + 1)
    )


def build_chart(ws, title, start, end):
    """A dual-axis chart over sheet rows ``start..end``: latency (col D) as
    columns on the left/Seconds axis, DRAM total (col E) as a line on the
    right/GB axis."""
    sheet = ws.title

    left = BarChart()
    left.type = "col"
    left.grouping = "clustered"
    left.gapWidth = 150
    left.roundedCorners = False
    left.title = title
    bar = Series(Reference(ws, min_col=LATENCY_COL, min_row=start, max_row=end))
    bar.graphicalProperties = GraphicalProperties(
        solidFill=LATENCY, ln=LineProperties(noFill=True)
    )
    bar.dLbls = _labels(LATENCY, pos="outEnd")
    left.series.append(_bind(bar, sheet, LATENCY_COL, start, end))

    right = LineChart()
    right.grouping = "standard"
    line = Series(Reference(ws, min_col=DRAM_COL, min_row=start, max_row=end))
    line.graphicalProperties = GraphicalProperties(
        ln=LineProperties(solidFill=DRAM, w=28575, cap="rnd")
    )
    line.marker = Marker(symbol="none")
    line.smooth = False
    line.dLbls = _labels(DRAM, pos="t")
    right.series.append(_bind(line, sheet, DRAM_COL, start, end))
    # Send the DRAM line to a secondary (right) axis; hide its category axis.
    right.y_axis.axId = 500
    right.x_axis.axId = 501
    right.y_axis.crosses = "max"
    right.x_axis.delete = True

    left += right

    # Title: 14pt Calibri, upright, its own band (overlay off).
    left.title.overlay = False
    left.title.tx.rich.p[0].pPr = ParagraphProperties(defRPr=_char(TITLE_PT))
    left.title.txPr = _text(TITLE_PT, rot=0)
    left.title.graphicalProperties = GraphicalProperties(
        noFill=True, ln=LineProperties(noFill=True)
    )

    # Category axis: hairline gray baseline, no ticks.
    x = left.x_axis
    x.delete = False
    x.majorTickMark = "none"
    x.minorTickMark = "none"
    x.tickLblPos = "nextTo"
    x.spPr = GraphicalProperties(
        noFill=True,
        ln=LineProperties(
            solidFill=GRID, w=9525, cap="flat", cmpd="sng", algn="ctr"
        ),
    )
    x.txPr = _text(CAT_PT)

    # Primary value axis (left): Seconds.  Start at 0, ~5 ticks, fixed decimals;
    # only this axis draws gridlines.  Extra headroom keeps the bars low.
    lat_max, lat_unit, lat_fmt = nice_axis(
        _col_max(ws, LATENCY_COL, start, end), headroom=LATENCY_HEADROOM
    )
    yl = left.y_axis
    yl.delete = False
    yl.axPos = "l"
    yl.majorTickMark = "none"
    yl.minorTickMark = "none"
    yl.tickLblPos = "nextTo"
    yl.majorGridlines = ChartLines(
        spPr=GraphicalProperties(
            ln=LineProperties(
                solidFill=GRID, w=9525, cap="flat", cmpd="sng", algn="ctr"
            )
        )
    )
    yl.scaling.min = 0
    yl.scaling.max = lat_max
    yl.majorUnit = lat_unit
    yl.numFmt = lat_fmt
    yl.spPr = GraphicalProperties(noFill=True, ln=LineProperties(noFill=True))
    yl.txPr = _text(VAL_PT)
    yl.dispUnits = _unit_axis("Seconds")

    # Secondary value axis (right): GB.  Its gridlines are off so they do not
    # draw a second (dark, default-colored) set of lines.
    dram_max, dram_unit, dram_fmt = nice_axis(
        _col_max(ws, DRAM_COL, start, end)
    )
    yr = right.y_axis
    yr.delete = False
    yr.axPos = "r"
    yr.majorTickMark = "none"
    yr.minorTickMark = "none"
    yr.tickLblPos = "nextTo"
    yr.majorGridlines = None
    yr.scaling.min = 0
    yr.scaling.max = dram_max
    yr.majorUnit = dram_unit
    yr.numFmt = dram_fmt
    yr.spPr = GraphicalProperties(noFill=True, ln=LineProperties(noFill=True))
    yr.txPr = _text(VAL_PT)
    yr.dispUnits = _unit_axis("GB")

    left.legend.position = "b"
    left.legend.overlay = False
    left.legend.txPr = _text(LEGEND_PT, rot=0)
    left.legend.spPr = GraphicalProperties(
        noFill=True, ln=LineProperties(noFill=True)
    )

    left.graphical_properties = GraphicalProperties(
        solidFill="FFFFFF", ln=LineProperties(noFill=True)
    )
    return left


def build_dram_breakdown(
    ws, title, start, end, series=BREAKDOWN_SERIES, unit_label="GB"
):
    """A stacked-bar DRAM-traffic breakdown over sheet rows ``start..end``.

    ``series`` is an ordered list of ``(legend name, metric field, RGB fill)``
    -- the default is Weight / Activation / KV Cache, matching the reference's
    chart5/6.  Pass a different list (e.g. Read / Write) for another stacked
    breakdown; ``unit_label`` names the value axis (all DRAM fields are bytes,
    so ``"GB"`` via dispUnits)."""
    sheet = ws.title
    cat = AxDataSource(
        strRef=StrRef(
            f=f"{sheet}!${_col(POINT_COL)}${start}:${_col(POINT_COL)}${end}"
        )
    )
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.gapWidth = BREAKDOWN_GAP
    chart.roundedCorners = False
    chart.title = title
    for name, field, fill in series:
        col = _field_col(field)
        s = Series(Reference(ws, min_col=col, min_row=start, max_row=end))
        s.tx = SeriesLabel(v=name)  # a literal legend name, not the header cell
        s.cat = cat
        s.graphicalProperties = GraphicalProperties(
            solidFill=fill, ln=LineProperties(noFill=True)
        )
        chart.series.append(s)

    # Title: 14pt Calibri, its own band -- set both the paragraph defRPr and the
    # chart txPr, or the title run falls back to a small default size.
    chart.title.overlay = False
    chart.title.tx.rich.p[0].pPr = ParagraphProperties(defRPr=_char(TITLE_PT))
    chart.title.txPr = _text(TITLE_PT, rot=0)
    chart.title.graphicalProperties = GraphicalProperties(
        noFill=True, ln=LineProperties(noFill=True)
    )

    # Category axis: hairline gray baseline, no ticks.
    x = chart.x_axis
    x.delete = False
    x.majorTickMark = "none"
    x.minorTickMark = "none"
    x.tickLblPos = "nextTo"
    x.spPr = GraphicalProperties(
        noFill=True,
        ln=LineProperties(
            solidFill=GRID, w=9525, cap="flat", cmpd="sng", algn="ctr"
        ),
    )
    x.txPr = _text(CAT_PT)

    # Value axis: light gridlines, GB via dispUnits, from 0 (Excel auto-picks
    # the max for the stacked total).
    y = chart.y_axis
    y.delete = False
    y.axPos = "l"
    y.majorTickMark = "none"
    y.minorTickMark = "none"
    y.tickLblPos = "nextTo"
    y.majorGridlines = ChartLines(
        spPr=GraphicalProperties(
            ln=LineProperties(
                solidFill=GRID, w=9525, cap="flat", cmpd="sng", algn="ctr"
            )
        )
    )
    y.scaling.min = 0
    y.numFmt = "#,##0"
    y.spPr = GraphicalProperties(noFill=True, ln=LineProperties(noFill=True))
    y.txPr = _text(VAL_PT)
    y.dispUnits = _unit_axis(unit_label)

    chart.legend.position = "b"
    chart.legend.overlay = False
    chart.legend.txPr = _text(LEGEND_PT, rot=0)
    chart.legend.spPr = GraphicalProperties(
        noFill=True, ln=LineProperties(noFill=True)
    )

    chart.graphical_properties = GraphicalProperties(
        solidFill="FFFFFF", ln=LineProperties(noFill=True)
    )
    return chart


def _band(g):
    """(from_row, to_row) chart band for group index ``g`` (0-based): a stack
    of 14-row-tall bands down the sheet."""
    if g == 0:
        return 0, 13
    return 13 + 14 * (g - 1), 13 + 14 * g


def _anchor(g, is_prefill, base_col=14):
    """Anchor placing group ``g``'s prefill chart on the left and decode on the
    right, in a band 14 rows tall.  ``base_col`` shifts the pair sideways so a
    second chart kind (the DRAM breakdown) sits to the right of the first."""
    from_row, to_row = _band(g)
    if is_prefill:
        frm = AnchorMarker(
            col=base_col, colOff=31308, row=from_row, rowOff=5721
        )
        to = AnchorMarker(
            col=base_col + 6, colOff=601494, row=to_row, rowOff=29813
        )
    else:
        frm = AnchorMarker(
            col=base_col + 6, colOff=552538, row=from_row, rowOff=5721
        )
        to = AnchorMarker(
            col=base_col + 13, colOff=449624, row=to_row, rowOff=29813
        )
    return TwoCellAnchor(_from=frm, to=to)


def _write_metric_sheet(ws, sheet):
    """Lay out one sweep's data (grouped: per group, prefill block then decode
    block, a blank row between groups) and drop a prefill/decode chart pair per
    group to the right."""
    ws.append(list(HEADER))
    # Explicit cursor: ws.max_row counts the last row *with data*, so it lags a
    # row right after a blank separator and would fold the blank into a range.
    row = 1
    blocks = []
    for g, group in enumerate(sheet.groups):
        for mode, is_pre in (("prefill", True), ("decode", False)):
            mrows = [r for r in group.rows if r["mode"] == mode]
            if not mrows:
                continue
            start = row + 1
            for r in mrows:
                ws.append(
                    [r.get("group", ""), r["point"], mode]
                    + [r.get(f) for f in METRIC_FIELDS]
                )
                row += 1
            end = row
            pfx = f"{sheet.prefix} " if sheet.prefix else ""
            head = f"{pfx}{mode.capitalize()}"
            blocks.append((g, is_pre, start, end, head, group.axis_title))
        ws.append([])
        row += 1
    # Each group gets two chart pairs side by side: latency+DRAM (left) and the
    # DRAM Weight/Activation/KV breakdown (right).
    for g, is_pre, start, end, head, axis in blocks:
        lat = build_chart(
            ws, f"{head} Latency and DRAM Traffic v.s. {axis}", start, end
        )
        lat.anchor = _anchor(g, is_pre, base_col=14)
        ws.add_chart(lat)
        brk = build_dram_breakdown(
            ws, f"{head} DRAM Traffic Breakdown v.s. {axis}", start, end
        )
        brk.anchor = _anchor(g, is_pre, base_col=28)
        ws.add_chart(brk)


def _fixup_label_numfmt(path):
    """openpyxl writes a data label's ``<numFmt>`` without ``sourceLinked``, so
    a reader ignores the format code and falls back to the cell's.  Stamp
    ``sourceLinked="0"`` on the ``#,##0.00`` label format so it takes effect."""
    pat = re.compile(r'(<numFmt formatCode="#,##0\.00")(\s*/?>)')
    buf = {}
    with zipfile.ZipFile(path) as z:
        names = z.namelist()
        for n in names:
            data = z.read(n)
            if n.startswith("xl/charts/chart") and n.endswith(".xml"):
                data = pat.sub(
                    r'\1 sourceLinked="0"\2', data.decode("utf-8")
                ).encode("utf-8")
            buf[n] = data
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
        for n in names:
            z.writestr(n, buf[n])


def write_aggregate(path, sheets):
    """Write one workbook with a named sheet per entry of ``sheets`` (each a
    :class:`MetricSheet` or :class:`BreakdownSheet`), then fix up the metric
    charts' label number formats."""
    wb = Workbook()
    wb.remove(wb.active)
    for sheet in sheets:
        ws = wb.create_sheet(sheet.name)
        if isinstance(sheet, MetricSheet):
            _write_metric_sheet(ws, sheet)
        else:
            per_module_latency_chart.populate_sheet(
                ws, sheet.rows, sheet.title, sheet.display_name
            )
    wb.save(path)
    _fixup_label_numfmt(path)
    return path
